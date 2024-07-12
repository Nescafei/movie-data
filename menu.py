from tkinter import *
from tkinter import messagebox, simpledialog
import tkinter as tk
from tkinter import ttk
from openpyxl import*
from tkvideo import tkvideo
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import os

#Excel
if os.path.exists('uwu.xlsx'):
    workbook = load_workbook('uwu.xlsx')
    ws = workbook['Sheet1']

else:
    workbook = Workbook()
    ws = workbook.active
    ws.title = 'Sheet1'

    columns = [
             "Movie Title",
             "Director",
             "Release Year",
             "Genre",
             "Availability",
             "Duration"
         ]

    ws.append(columns)
    workbook.save('uwu.xlsx')

#First Interface
root = tk.Tk()
root.geometry('1200x600')
root.title('NesCinema')
root.configure(bg='#BFACAB')
root.tk.call('wm', 'iconphoto', root._w, tk.PhotoImage(file='Iconsz.png'))
root.resizable(FALSE, FALSE)

wb = Workbook
wb = load_workbook('uwu.xlsx')
ws = wb.active

#Photo Image Function
pic = PhotoImage(file="nesbg.png")
pic2 = PhotoImage(file="ent.png")
pic3 = PhotoImage(file="ext.png")
my =  Label(root, image=pic).place(x=0, y=0)
my2 =  Label(root, image=pic2)
my3 =  Label(root, image=pic3)
lab = Label(root, text="")

#Video Interface
lbl = Label(root)
lbl.place(x=0, y=0)

x = tkvideo("vidbg.mp4",
            lbl,
            loop=1,
            size=(1200,600))
x.play()


def eq():
    root.withdraw()


#Main Interface
def new():
    miw = tk.Toplevel()
    miw.title('NesCinema Main Interface')
    miw.geometry('1350x700')
    miw.resizable(FALSE, FALSE)
    miw.configure(bg='#5C3847')
    root.withdraw()

#Main Interface Frames
    f1 = Frame(miw, width=1330, height=65, bg="#F5D1B1", bd="2", highlightthickness='3', highlightbackground="#C9A280")
    f1.grid(row=0, column=0, pady=8, padx=9)
    f2 = Frame(miw, width=660, height=500, bg="#F5D1B1", highlightthickness='3', highlightbackground="#C9A280")
    f2.grid(row=1, column=0, padx=9, sticky="w")

    f3 = Frame(miw, width=660, height=246, bg="#F5D1B1", highlightthickness='3', highlightbackground="#C9A280")
    f3.place(x=679, y=81)

    f4 = Frame(miw, width=660, height=246, bg="#F5D1B1", highlightthickness='3', highlightbackground="#C9A280")
    f4.place(x=679, y=335)
    f5 = Frame(miw, width=1330, height=100, bg="#F5D1B1", bd="6", highlightthickness='3', highlightbackground="#C9A280")
    f5.grid(row=2, column=0, pady=8, padx=9)

#Main Interface Labels
    lb1 = Label(miw, text="NesCinema Management System", bg="#F5D1B1", fg="#664C35", font=("Comic Sans MS", 27, "bold")).place(x=440, y=12)
    lb2 = Label(miw, text="Movie Title", bg="#F5D1B1", fg="#664C35", font=("Lithos Pro Regular", 15, "bold")).place(x=15, y=110)
    lb3 = Label(miw, text="Director", bg="#F5D1B1", fg="#664C35", font=("Lithos Pro Regular", 15, "bold")).place(x=15, y=170)
    lb4 = Label(miw, text="Genre", bg="#F5D1B1", fg="#664C35", font=("Lithos Pro Regular", 15, "bold")).place(x=15, y=230)
    lb5 = Label(miw, text="Casts", bg="#F5D1B1", fg="#664C35", font=("Lithos Pro Regular", 15, "bold")).place(x=15, y=290)
    lb6 = Label(miw, text="Release Year", bg="#F5D1B1", fg="#664C35", font=("Lithos Pro Regular", 15, "bold")).place(x=15, y=350)
    lb7 = Label(miw, text="Availability", bg="#F5D1B1", fg="#664C35", font=("Lithos Pro Regular", 15, "bold")).place(x=15, y=410)
    lb8 = Label(miw, text="Duration", bg="#F5D1B1", fg="#664C35", font=("Lithos Pro Regular", 15, "bold")).place(x=15, y=470)
    lb9 = Label(miw, text="Result", bg="#F5D1B1", fg="#664C35", font=("Comic Sans MS", 15, "bold")).place(x=1000, y=340)
    lb10 = Label(miw, text="Title: ", bg="#F5D1B1", fg="#664C35", font=("Lithos Pro Regular", 13, "bold")).place(x=690, y=380)
    lb11 = Label(miw, text="Director: ", bg="#F5D1B1", fg="#664C35", font=("Lithos Pro Regular", 13, "bold")).place(x=690, y=405)
    lb12 = Label(miw, text="Genre: ", bg="#F5D1B1", fg="#664C35", font=("Lithos Pro Regular", 13, "bold")).place(x=690, y=430)
    lb13 = Label(miw, text="Cast: ", bg="#F5D1B1", fg="#664C35", font=("Lithos Pro Regular", 13, "bold")).place(x=690, y=455)
    lb14 = Label(miw, text="Release Year: ", bg="#F5D1B1", fg="#664C35", font=("Lithos Pro Regular", 13, "bold")).place(x=690, y=480)
    lb15 = Label(miw, text="Availability: ", bg="#F5D1B1", fg="#664C35", font=("Lithos Pro Regular", 13, "bold")).place(x=690, y=505)
    lb16 = Label(miw, text="Duration: ", bg="#F5D1B1", fg="#664C35", font=("Lithos Pro Regular", 13, "bold")).place(x=690, y=530)

#Main Interface Entries
    movie_title = Entry(miw, width=40, bg="#F9E5D3", bd="4", font=("Lithos Pro Regular", 15, "bold"))
    movie_title.place(x=160, y=110)
    director = Entry(miw, width=40, bg="#F9E5D3", bd="4", font=("Lithos Pro Regular", 15, "bold"))
    director.place(x=160, y=170)
    cast = Entry(miw, width=40, bg="#F9E5D3", bd="4", font=("Lithos Pro Regular", 15, "bold"))
    cast.place(x=160, y=290)
    year_rel = Entry(miw, width=40, bg="#F9E5D3", bd="4", font=("Lithos Pro Regular", 15, "bold"))
    year_rel.place(x=160, y=350)
    duration = Entry(miw, width=40, bg="#F9E5D3", bd="4", font=("Lithos Pro Regular", 15, "bold"))
    duration.place(x=160, y=470)

#Main Interface ComboBox
    genvar = StringVar()
    genlist = ['Comedy', 'Horror', 'Action', 'Fantasy', 'Mystery', 'Drama', 'Adventure', 'Documentary', 'Sci-Fi']
    gencombo = ttk.Combobox(miw, width=39, font=("Lithos Pro Regular", 15, "bold"), value=genlist, textvariable=genvar)
    gencombo.place(x=160, y=230)

    availvar = StringVar()
    avail_list = ['CD', 'Netflix', 'Cinemas', 'HBO', 'Disney+']
    avail_combo = ttk.Combobox(miw, width=39, font=("Lithos Pro Regular", 15, "bold"), value=avail_list, textvariable=availvar)
    avail_combo.place(x=160, y=410)

#Main Interface Functions 

    #Global
    def glob():
        global mov, dir, c, gen, av, dur, yr
        mov = movie_title.get()
        dir = director.get()
        c = cast.get()
        yr = year_rel.get()
        gen = genvar.get()
        av = availvar.get()
        dur = duration.get()

    #Add/Save Function
    def create():
        mov = movie_title.get()
        dir = director.get()
        c = cast.get()
        yr = year_rel.get()
        gen = genvar.get()
        av = availvar.get()
        dur = duration.get()
        Found = False
        for ec in range(2,(ws.max_row)+1):
            if(mov == ws['A'+str(ec)] or (dir == ws['B'+str(ec)] or (c == ws['C'+str(ec)] or (yr == ws['D'+str(ec)] or (gen == ws['E'+str(ec)] or (av == ws['F'+str(ec)] or (dur == ws['G'+str(ec)].value))))))):
                Found = True
                break
            else:
                Found = False
                
        if(Found == True):
            messagebox.showerror("Error", "Data Already Exist!")
        else:
            lastrow = str(ws.max_row+1)
            ws['A'+ lastrow] = movie_title.get()
            ws['B'+ lastrow] = director.get()
            ws['C'+ lastrow] = cast.get()
            ws['D'+ lastrow] = year_rel.get()
            ws['E'+ lastrow] = genvar.get()
            ws['F'+ lastrow] = availvar.get()
            ws['G'+ lastrow] = duration.get()
            reset()
            messagebox.showinfo("Noice", "Data Saved Successfully!")
        wb.save('uwu.xlsx')
        refresh(t1)


    def show(cell):
        reset()
        movie_title.insert(0,ws['A'+str(cell + 1)].value)
        director.insert(0,ws['B'+str(cell + 1)].value)
        cast.insert(0,ws['C'+str(cell + 1)].value)
        year_rel.insert(0,ws['D'+str(cell + 1)].value)
        gencombo.set(ws['E'+str(cell + 1)].value)
        avail_combo.set(ws['F'+str(cell + 1)].value)
        duration.insert(0,ws['G'+str(cell + 1)].value)
        glob()



    #Search Function
    def search():
        nr = Toplevel()
        nr.geometry('300x150')
        nr.title("Search")
        nr.configure(bg='#F5D1B1')
        nr.tk.call('wm', 'iconphoto', nr._w, tk.PhotoImage(file='Iconsz.png'))
        nr.resizable(FALSE,FALSE)
        py = PhotoImage(file='searchbg.png')
        pi = Label(nr, image=py).place(x=0, y=0)


        def searc():
            ref = []
            Found = False
            for ec in range(2, (ws.max_row)+1):
                if(sr1.get() == ws['A'+str(ec)].value):
                    Found = True
                    ref.append(ec-1)
                    cell_address = str(ec-1)
                    break

            if(Found == True):
                messagebox.showinfo("Data Found", "Data Exist In Cell " + cell_address)
                Found = False
                show(ref[0])

            else:
                messagebox.showerror("Error 404", "Title Not Found")



        sr = Label(nr, text="Search Title", fg="#664C35", font=("Comic Sans MS", 14, "bold"), bg="#BFACAB").place(x=90,y=5)
        sr1 = Entry(nr, width=38, bg="#F9E5D3", bd="3")
        sr1.place(x=35, y=40)

        def ex():
            nr.withdraw()

        btn1 = Button(nr, text="Search", width=30, bd="3", fg="#3C1E09", bg="#EAD0CE", font=("Comic Sans MS", 8, "bold"), command=lambda:searc()).place(x=40, y=75)
        btn1 = Button(nr, text="Exit", width=30, bd="3", fg="#3C1E09",bg="#EAD0CE", font=("Comic Sans MS", 8, "bold"),command=lambda:ex()).place(x=40, y=110)
        nr.mainloop()


    #Update Function
    def update():
        Found = True
        for ec in range(2, (ws.max_row)+1):
            if (mov ==  ws['A'+str(ec)].value):
                cell_address = ec
                Found = True
                break
            else:
                Found = False
        if (Found == True):
            ws['A'+str(cell_address)].value = movie_title.get()
            ws['B'+str(cell_address)].value = director.get()
            ws['C'+str(cell_address)].value = cast.get()
            ws['D'+str(cell_address)].value = year_rel.get()
            ws['E'+str(cell_address)].value = genvar.get()
            ws['F'+str(cell_address)].value = availvar.get()
            ws['G'+str(cell_address)].value = duration.get()
            messagebox.showinfo("Updated!", "Movie Info Has Been Updated")
            reset()
        workbook.save('uwu.xlsx')
        refresh(t1)



    #Delete/Remove Data Function
    def delete():
        for ec in range(2, (ws.max_row)+1):
            if(movie_title.get() == ws['A'+str(ec)] or (director.get() == ws['B'+str(ec)] or (cast.get() == ws['C'+str(ec)] or (year_rel.get() == ws['D'+str(ec)] or (genvar.get() == ws['E'+str(ec)] or (availvar.get() == ws['F'+str(ec)] or (duration.get() == ws['G'+str(ec)].value))))))):
                Found = True
                cell_address = ec
                break
            else:
                Found = False

        if(Found == True):
            ws.delete_rows(ec)
            messagebox.showinfo("Data Yeeted", "Data Removed")
            reset()
        wb.save('uwu.xlsx')
        refresh(t1)

        if(Found == False):
                    messagebox.showerror("Error", "Please a valid input")
                    reset()


    #Reset Entries Function
    def reset():
        mov = movie_title.delete(0, END)
        dir = director.delete(0, END)
        c = cast.delete(0, END)
        yr = year_rel.delete(0, END)
        gen = genvar.set("")
        av = availvar.set("")
        dur = duration.delete(0, END)


    #Refresh Function
    def restart():
        yesno = messagebox.askquestion("Restart", "Your data will be lost, are you sure?", icon='warning')
        if yesno == 'yes':
            miw.withdraw()
            new()

    def refresh(tree):
        tree.delete(*tree.get_children())
        data = updateddata()
        for item in data:
            tree.insert('', 'end', values=item)
    def updateddata():
        updated_value = list()
        for each_cell in range(2, (ws.max_row)+1):     
            updated_value.append([ws['A'+str(each_cell)].value, ws['B'+str(each_cell)].value, ws['D'+str(each_cell)].value, ws['E'+str(each_cell)].value, ws['F'+str(each_cell)].value, ws['G'+str(each_cell)].value, ws['C'+str(each_cell)].value])
        return updated_value

    #Show Data Function
    def result():
        global result1, result2, result3, result4, result5, result6, result7
        result1 = Label(miw, bg="#F5D1B1", fg="#1E160D", font=("Lithos Pro Regular", 12, "bold"), text=f'{movie_title.get()}').place(x=810, y=380)
        result2 = Label(miw, bg="#F5D1B1", fg="#1E160D", font=("Lithos Pro Regular", 12, "bold"), text=f'{director.get()}').place(x=810, y=405)
        result3 = Label(miw, bg="#F5D1B1", fg="#1E160D", font=("Lithos Pro Regular", 12, "bold"), text=f'{genvar.get()}').place(x=810, y=430)
        result4 = Label(miw, bg="#F5D1B1", fg="#1E160D", font=("Lithos Pro Regular", 12, "bold"), text=f'{cast.get()}').place(x=810, y=455)
        result5 = Label(miw, bg="#F5D1B1", fg="#1E160D", font=("Lithos Pro Regular", 12, "bold"), text=f'{year_rel.get()}').place(x=810, y=480)
        result6 = Label(miw, bg="#F5D1B1", fg="#1E160D", font=("Lithos Pro Regular", 12, "bold"), text=f'{availvar.get()}').place(x=810, y=505)
        result7 = Label(miw, bg="#F5D1B1", fg="#1E160D", font=("Lithos Pro Regular", 12, "bold"), text=f'{duration.get()}').place(x=810, y=530)

    def help():
        helpz = tk.Toplevel()
        helpz.geometry('1200x550')
        helpz.title("Help")
        helpz.configure(bg="#BFACAB")
        helpz.tk.call('wm', 'iconphoto', helpz._w, tk.PhotoImage(file='Iconsz.png'))
        helpz.resizable(FALSE, FALSE)
        lbl2 = Label(helpz)
        lbl2.place(x=0, y=0)

        x2 = tkvideo("Golden Hour(2).mp4",
                    lbl2,
                    loop=1,
                    size=(1200,550))

        x2.play()


        def exitz():
            helpz.withdraw()

        def refreshss():
            helpz.withdraw()
            help()

        h_title = Label(helpz, text="Guidelines", bg="#AB545D", font=("Comic Sans MS", 16, "bold")).pack()
        h_descreate = Label(helpz, bg="#AB545D", font=("Arial", 11, "bold"), text="Press to add a certain data that has been put in the entry.\n Once the user decides to add it, the data will pop in the treeview.").place(x=175,y=90)
        h_ref = Label(helpz, bg="#AB545D", font=("Arial", 11, "bold"), text="Restarts the window entirely.").place(x=175, y=200)
        h_updates = Label(helpz, bg="#AB545D", font=("Arial", 11, "bold"), text="The user can manipulate the data within the interface.\n Note that this function only works \nwhen you SEARCH the specific title to edit." ).place(x=175, y=310)
        h_del = Label(helpz, bg="#AB545D", font=("Arial", 11, "bold"), text="Removes a certain data inside the Treeview within the Excel file.").place(x=175, y=420)
        h_searchz = Label(helpz, bg="#AB545D", font=("Arial", 11, "bold"), text="Pops up a new interface that can able to search\n the data's location inside the Excel file.").place(x=795, y=90)
        h_resetz = Label(helpz, bg="#AB545D", font=("Arial", 11, "bold"), text="Wipes all the data inside the entries.").place(x=795, y=200)
        h_showz = Label(helpz, bg="#AB545D", font=("Arial", 11, "bold"), text="Shows the result of the entries at the lower right frame.").place(x=795, y=310)
        h_exit = Label(helpz, bg="#AB545D", font=("Arial", 11, "bold"), text="Exits the interface.").place(x=795, y=420)

        h_create = Button(helpz, text="Create", width=10, bd=3, bg="#EAD0CE", font=("Comic Sans MS", 13, "bold")).place(x=40, y=90)
        h_refresh = Button(helpz, text="Refresh", width=10, bd=3, bg="#EAD0CE", font=("Comic Sans MS", 13, "bold"), command=lambda:refreshss()).place(x=40, y=200)
        h_update = Button(helpz, text="Update", width=10, bd=3, bg="#EAD0CE", font=("Comic Sans MS", 13, "bold")).place(x=40, y=310)
        h_delete = Button(helpz, text="Delete", width=10, bd=3, bg="#EAD0CE", font=("Comic Sans MS", 13, "bold")).place(x=40, y=420)
        h_search = Button(helpz, text="Search", width=10, bd=3, bg="#EAD0CE", font=("Comic Sans MS", 13, "bold")).place(x=655, y=90)
        h_reset = Button(helpz, text="Reset Entry", width=10, bd=3, bg="#EAD0CE", font=("Comic Sans MS", 13, "bold")).place(x=655, y=200)
        h_show = Button(helpz, text="Show Data", width=10, bd=3, bg="#EAD0CE", font=("Comic Sans MS", 13, "bold")).place(x=655, y=310)
        h_exit = Button(helpz, text="Exit", width=10, bd=3, bg="#EAD0CE", font=("Comic Sans MS", 13, "bold"), command=lambda:exitz()).place(x=655, y=420)
        helpz.mainloop()

    #Exit/Quit Function
    def exit():
        yesno = messagebox.askquestion("Confirmation", "Your data will be lost, are you sure?", icon='warning')
        if yesno == 'yes':
            miw.withdraw()

    #Main Interface Buttons
    bt1 = Button(miw, text="Create", width=10, bd="5", fg="#3C1E09", bg="#FFC591",font=("Comic Sans MS", 15, "bold"), command=lambda:create()).place(x=25, y=610)
    bt2 = Button(miw, text="Refresh", width=10, bd="5", fg="#3C1E09", bg="#FFC591",font=("Comic Sans MS", 15, "bold"), command=lambda:restart()).place(x=190, y=610)
    bt3 = Button(miw, text="Update", width=10, bd="5", fg="#3C1E09", bg="#FFC591",font=("Comic Sans MS", 15, "bold"), command=lambda:update()).place(x=355, y=610)
    bt4 = Button(miw, text="Delete", width=10, bd="5", fg="#3C1E09", bg="#FFC591",font=("Comic Sans MS", 15, "bold"), command=lambda:delete()).place(x=520, y=610)
    bt5 = Button(miw, text="Search", width=10, bd="5", fg="#3C1E09", bg="#FFC591",font=("Comic Sans MS", 15, "bold"), command=lambda:search()).place(x=685, y=610)
    bt6 = Button(miw, text="Show Data", width=10, bd="5", fg="#3C1E09", bg="#FFC591",font=("Comic Sans MS", 15, "bold"), command=lambda:result()).place(x=1025, y=610)
    bt7 = Button(miw, text="Reset Entry", width=10, bd="5", fg="#3C1E09", bg="#FFC591",font=("Comic Sans MS", 15, "bold"), command=lambda:reset()).place(x=855, y=610)
    bt8 = Button(miw, text="Exit", width=10, bd="5", fg="#3C1E09", bg="#FFC591",font=("Comic Sans MS", 15, "bold"), command=lambda:exit()).place(x=1190, y=610)
    bt9 = Button(miw, text="?", width=3, bd="5", fg="#3C1E09", bg="#FFC591",font=("Arial", 10, "bold"), command=lambda:help()).place(x=1290, y=538)

    #TreeView Function (Useless Scrollbar)
    def view():
        global t1
        t1 = ttk.Treeview(f3, height=10)
        
        treescrolly = Scrollbar(f3, orient="vertical", command=t1.yview)
        treescrollx = Scrollbar(f3, orient="horizontal", command=t1.xview)
        t1.configure(xscrollcommand = treescrollx.set,yscrollcommand=treescrolly.set)
        treescrollx.pack(side ="bottom",fill ="x")
        treescrolly.pack(side ="right",fill="y")  

        t1['columns'] = ("Title", "Director", "Release Year", "Genre", "Availability", "Duration")
        t1.column("Title", anchor=W, width=160)
        t1.column("Director",  anchor=W, width=130)
        t1.column("Release Year", anchor=CENTER, width=80)
        t1.column("Genre", anchor=CENTER, width=80)
        t1.column("Availability", anchor=CENTER, width=80)
        t1.column("Duration", anchor=CENTER, width=105)
        t1["show"]="headings"
        t1.pack(fill=BOTH,expand=1)

        t1.heading("Title", text="Title")
        t1.heading("Director", text="Director")
        t1.heading("Release Year", text="Release Year")
        t1.heading("Genre", text="Genre")
        t1.heading("Availability", text="Availability")
        t1.heading("Duration", text="Duration")
        
        
        for ec in range(2, (ws.max_row)+1):
 
            t1.insert(parent='', index="end", values=(ws['A'+str(ec)].value, ws['B'+str(ec)].value, ws['D'+str(ec)].value, ws['E'+str(ec)].value, ws['F'+str(ec)].value, ws['G'+str(ec)].value))
        t1.pack()
    view()

column_A = ws['A']
column_B = ws['B']
column_D = ws['D']
column_E = ws['E']
column_F = ws['F']
column_G = ws['G']


#Login Interface Buttons
enter_btn = Button(root, image=pic2,bg='#5C3847', borderwidth=0, highlightthickness=0, relief="flat", command=lambda:new())
enter_btn.place(x=830, y=160)

exit_btn = Button(root, image=pic3,bg='#5C3847', borderwidth=0, highlightthickness=0, relief="flat", command=lambda:eq())
exit_btn.place(x=830, y=325)

root.mainloop()